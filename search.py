from dotenv import load_dotenv
from job_scraper_utils import *
import json
from openai import OpenAI
from datetime import datetime
from docx import Document
from docx.shared import Pt
from docx.oxml.ns import qn
from docx.oxml import OxmlElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from openpyxl import Workbook
from openpyxl import load_workbook
import re
from ast import literal_eval

load_dotenv()
api_key = os.getenv('OPENAI_API_KEY')
united_states = 'https://www.indeed.com'
country = united_states

driver = configure_webdriver()
# job_positions = ['finance analyst', 'business operations', 'strategic finance', 'financial associate', 'technology finance', 'corporate development']
# job_positions = ['finance analyst', 'financial analyst', 'strategic finance', 'finance associate']
job_positions = ['Strategic finance']
# job_location = '90066'
# job_locations = ['Los Angeles', 'New York City', 'San Francisco']
job_locations = ['Los Angeles']
date_posted = 1
pay = '$130,000'
yearsExperience = '4'
disqualifySkills = 'Director, partner, lead, VP, Vice President, Manager, Equity Research, underwriting'
disqualifyTerms = 'Director, partner, lead, VP, Vice President, Manager, Equity Research, underwriting'

base_resume = """
A. Silicon Valley Bank, San Francisco, CA | Senior Associate   
1. •	Led strategic finance initiatives for new transactions of exceeding $50+ million to drive decision making by building liquidity analysis, Excel forecasting models, drafting term sheets to clients, and underwriting the final memo for enterprise software companies that raised Series A through D equity rounds
2. •	Collaborated with cross-functional teams to design and implement growth strategies for high-value enterprise software clients, driving $150+ million in debt transactions and expanding the debt portfolio by 25%
3. •	Developed and implemented an Excel-based portfolio tracking system, streamlining KPI reporting and enabling accurate valuation analyses, which improved operational efficiency and data-driven decision-making by 200%
4. •	Trained and supervised 12+ junior associates in financial accounting, underwriting, portfolio management, and due diligence responsibilities, serving as the go-to resource for ad hoc inquiries 
5. •	Spearheaded due diligence and investment strategies for 30+ B2B technology clients, presenting detailed business updates and strategic recommendations to senior leadership to enable data-driven decision-making
B. Bridge Bank, San Francisco, CA | Technology Lending Analyst
6. •	Assisted in developing innovative debt structures such as venture debt loans and recurring revenue LOC, closing over 15+ deals with total loan commitments exceeding $100+ million, demonstrating strategic financial planning 
7. •	Streamlined a portfolio management process for 20+ companies by optimizing the collection and analysis of monthly reports, aligning company performances with strategic growth plans 
8. •	Engaged in due diligence discussions with company executives and investors—covering venture capital, private equity, and strategic investors—to develop capital structure solutions for technology and innovation sectors, honing KPI reporting and financial analysis skills
C. Wells Fargo & Company, Omaha, NE | Middle Market Banking Analyst
9. •	Attended a six-month analyst program structured around learning financial modeling, and credit underwriting 
10. •	Initiated transactions and managed a portfolio of 10+ companies with revenues from $5 million to $2 billion from various industries including investor real estate, private equity, telecommunications, and agriculture 
11. •	Partnered with relationship teams to structure and propose credit transactions as well as coordinate customer oversight, documentation, and loan closing processes for new and existing customers 
"""


# This function creates a search based on the variables above and outputs to a json file.
def searchToJson():
    today_date = datetime.today().strftime('%Y-%m-%d')
    filename = f"{today_date}/{today_date}.json"
    dataframes = []
    for job_location in job_locations:
        for job_position in job_positions:
            search_jobs(pay, driver, country, job_position, job_location, date_posted)
            df = scrape_job_data(driver, country)
            dataframes.append(df)
    # Concatenate all dataframes and remove duplicates
    merged_df = pd.concat(dataframes).drop_duplicates().reset_index(drop=True)
    merged_df.to_json(filename, orient='records')
    json_data = merged_df.to_json(orient="records")
    # Parse and pretty-print JSON
    parsed_json = json.loads(json_data)
    with open(filename, "w") as f:
        json.dump(parsed_json, f, indent=4)


# Given a string, return a list of the only index we want to keep. VERY STRICT WITH THE FORMAT.
def addGoodJobs(partialPrompt):
    print('Adding only best job')
    prompt = f"""
            Given the following details:
            Base resume: {base_resume}
            List of job descriptions with indices: {partialPrompt}
            Years of experience threshold: {yearsExperience}
            Disqualify terms: {disqualifyTerms}

        Instructions:
        For each job description:

            If the estimated years of experience exceed {yearsExperience}, mark the job as DISQUALIFIED.
            If the job is not a good fit for the resume, mark the job as DISQUALIFIED.
            If the job title or description contains any of the disqualify terms, mark the job as DISQUALIFIED.
            If the job is not strictly a finance role (FP&A, finance analyst/strategy), mark the job as DISQUALIFIED.
            If the company is not tech-focused, mark the job as DISQUALIFIED.

            If none of the jobs are qualified, output : []
            
            Otherwise output the job with best fit in a python array like this : [13]

            Do not include any explanations, additional text, or formatting beyond the specified output format.
            """

    #     Final Output:

    #         If a job is DISQUALIFIED for ANY reason, exclude it from consideration.
    #         From the remaining jobs, choose the most relevant job based on role similarity (key skills, responsibilities, and title overlap).
    #         If no jobs are suitable, return an empty list.

    #     Output format:

    #         Return only the index of the most relevant job in a Python list (e.g., [13]).
    #         If no jobs qualify, return an empty list (e.g., []).
    #         Do not include any explanations, additional text, or formatting beyond the specified output format.
    # """
    client = OpenAI(api_key=api_key)
    response = client.chat.completions.create(
        model="gpt-4",
        messages=[
            {"role": "system", "content": "You are an assistant that returns strictly structured outputs based on user instructions."},
            {"role": "user", "content": prompt}
        ],
    )
    print('addGoodJobs prompt response = ', response.choices[0].message.content)
    stringOutput = response.choices[0].message.content.strip()  # Remove any surrounding whitespace
    print('Index we are adding =', stringOutput)
    # Check if the string output contains a valid list format
    try:
        # Use Python's eval safely with literal_eval to parse the string as a Python list
        parsed_output = literal_eval(stringOutput)
        # Ensure the parsed output is a list of integers
        if isinstance(parsed_output, list) and all(isinstance(i, int) for i in parsed_output):
            int_list = parsed_output
        else:
            # If the parsed output isn't a valid list of integers, return an empty list
            int_list = []
    except (ValueError, SyntaxError):
        # If parsing fails, return an empty list
        int_list = []
    print('Job index we are adding = ', int_list)
    return int_list

# given a job description, this function summarizes it. 
def summarize(jobDescription):
    print('summarizing job description...')
    prompt = f"""
        Below is a job description. Summarize it in no more than 200 words by identifying:
        1. Job Title
        2. Primary technologies, tools, and skills required
        3. Years of experience required
        4. Key responsibilities
        5. Whether it specifies industry knowledge (e.g., finance, healthcare)

        Job description:
        {jobDescription}
    """
    client = OpenAI(api_key=api_key)
    response = client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=[
            {"role": "system", "content": "You are a technical recruiter who specializes in financial jobs."},
            {"role": "user", "content": prompt}
        ],
    )
    return response.choices[0].message.content

# Given a dict : key = company, value = [index, description], return a list of index to keep.
# Take the unique company names. Then go through dictionary and merge the index and company name in a readable string.
# Then invoke openAI to return only the best one.
def addingIndexes(descriptionDict):
    if len(descriptionDict) == 0:
        return set()
    result = set()
    for company in descriptionDict:
        partialPrompt = ''
        for job in range(len(descriptionDict[company])):
            partialPrompt += 'INDEX = ' + str(descriptionDict[company][job][0]) + '|' + summarize(descriptionDict[company][job][1])
        print('working on duplicates for company ', company)
        result.update(addGoodJobs(partialPrompt))
    return result

# We want a refine stage between the job scrape and the resume building process. 
# Here is where you put all the conditions you want to include and make sure you don't apply to a job.
# Initially doing this to ensure we only apply to one job per company per day. We need to select the best one. 
# This returns a list of index we are keeping (length = 0 or 1).
# Increasing this now to include jobs that are too senior or include keywords that don't match. This could be
# done in one prompt. 
# I am going to pass in the entire job description dict : key = company, value = index, summarized description.

def refineJson(filename):
    companiesDescriptions = {}
    added = set()
    with open(filename, "r") as f:
        data = json.load(f)
        index = 0
        for entry in data:
            company = entry.get("Company")
            url = entry.get("Link")
            if company not in companiesDescriptions:
                companiesDescriptions[company] = []
            companiesDescriptions[company].append([index, jobDescription(url)])
            index += 1
        print('Filtering jobs')
        toAdd = set()
        toAdd = addingIndexes(companiesDescriptions)
        for r in toAdd:
            added.add(r)
    return added


# This function takes a URL and returns the full description as a string.
def jobDescription(url):
    if not url:
        return 'ERROR, No Job Description!'
    try:
        driver.get(url)
        description_element = driver.find_element(By.XPATH, '//div[contains(@class, "jobsearch-JobComponent-description")]')
        description_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//div[contains(@class, "jobsearch-JobComponent-description")]'))
            )
        description = description_element.text
        start_index = description.find("Full job description")
        if start_index != -1:
            description = description[start_index:]
        return description
    except (TimeoutException, NoSuchElementException):
        print("Job description element not found or page took too long to load.")
        return None


def generate_resume_bullets(job_description):
    prompt = f"""
        Here is a partial resume with numbered bullet points : {base_resume}
        And here is a job description {job_description}
        
        STEP 1 (do not output anything) : 
        Find the keywords for experience/technology/skills that are present in the job description and not present in 
        the resume. If none, output 'NO CHANGES'.

        STEP 2 (do not output anything):
        Take the keywords/skills from step 1 and rank them in order of importance in relation to the job. 

        STEP 3 (do not output anything):
        Create 1-4 bullets: 
        - Do not include anything that cannot be justified within the resume. Keep it at least somewhat related.
        - Write in a consice and precise manner. Minimize fluff. Use clear terms. Make it shorter or equal length as the average bullet in the resume.
        - Include as many missing keywords and skills as possible in each bullet.
        - Make sure each bullet is formatted like the bullets in the resume (in one cohesive manner) : 
            'Improved X by Y through implementation of Z.'
        - Use arbitrary metrics that sound very impressive : Accomplishments, Percentages, milestones, requests etc...
        
        STEP 4 (do not output anything):
        Find which Section (labeled A-C) is most relevant for each bullet you've created.

        STEP 5 (do not output anything):
        Within each Section (from step 4), save the bullet number(s) that is/are least relevant to the job.
        Each new created bullet (from step 3) must have a unique bullet number it is replacing.
        Add that number at the end of the new bullets you've created. Just the number. No 'Replaces bullet'.
        Remove any number at the start of the bullet output. Only keep the sentence:
        For example : 
        Increased scalability and reliability of 6+ applications by 40%, migrating hybrid infrastructure to AWS ECS, ECR, RDS, and MQ with Terraform. 1
        Remove all quotes from the bullet points. 

        STEP 6 (OUTPUT):
        Output each bullet from step 5 followed by a number.

    """
    client = OpenAI(api_key=api_key)
    response = client.chat.completions.create(
        model="gpt-4",
        messages=[
            {"role": "system", "content": "You are a very sophisticated computer program. Each instruction is detailed in steps that you will follow perfectly. Each step will follow this output 'STEP X:' followed by the expected and predictable output of that step. Never skip lines. Explanations are forbidden. "},
            {"role": "user", "content": prompt}
        ],
        max_tokens=1000,
        temperature=0.2,
    )
    return response.choices[0].message.content

# We get an output of a couple bullet points with numbers. Now we want to replace the original bullets with the new ones
# Bullet numbers will need to be hardcoded. 
# This function will take the GPT output and return a string including only the bullets
def isolateBullets(GPTOutput):
    start_index = GPTOutput.find("STEP 6:")
    if start_index != -1:
        result = GPTOutput[start_index + len("STEP 6:"):].strip()
        return result
    else:
        print("STEP 6: not found in the input.")
        return 'ERROR'
    
# From the gpt output, we create a dictonary where key = bullet to replace and value = new bullet. 
def parse_string_to_dict(input_string):
    lines = input_string.splitlines()
    result_dict = {}
    for line in lines:
        # Split each line by the last space to separate the text from the index
        *text, index = line.rsplit(" ", 1)
        try:
            result_dict[int(index)] = " ".join(text)
        except:
            print('ERROR with prompt : ', input_string)
    return result_dict

# This replaces the bullets and creates a new updated resume
# doc_path is the output path, and newBulletDict is the bullet dict where key = bullet # and value = new bullet
def replaceBullets(doc_path, newBulletDict):
    doc = Document(docx_path)
    # Define the bullet points by number in a dictionary
    bullet_points = {
        1: 'Led strategic finance initiatives for new transactions of exceeding $50+ million to drive decision making by building liquidity analysis, Excel forecasting models, drafting term sheets to clients, and underwriting the final memo for enterprise software companies that raised Series A through D equity rounds',
        2: 'Collaborated with cross-functional teams to design and implement growth strategies for high-value enterprise software clients, driving $150+ million in debt transactions and expanding the debt portfolio by 25%',
        3: 'Developed and implemented an Excel-based portfolio tracking system, streamlining KPI reporting and enabling accurate valuation analyses, which improved operational efficiency and data-driven decision-making by 200%',
        4: 'Trained and supervised 12+ junior associates in financial accounting, underwriting, portfolio management, and due diligence responsibilities, serving as the go-to resource for ad hoc inquiries',
        5: 'Spearheaded due diligence and investment strategies for 30+ B2B technology clients, presenting detailed business updates and strategic recommendations to senior leadership to enable data-driven decision-making',
        6: 'Assisted in developing innovative debt structures such as venture debt loans and recurring revenue LOC, closing over 15+ deals with total loan commitments exceeding $100+ million, demonstrating strategic financial planning',
        7: 'Streamlined a portfolio management process for 20+ companies by optimizing the collection and analysis of monthly reports, aligning company performances with strategic growth plans',
        8: 'Engaged in due diligence discussions with company executives and investors—covering venture capital, private equity, and strategic investors—to develop capital structure solutions for technology and innovation sectors, honing KPI reporting and financial analysis skills',
        9: 'Attended a six-month analyst program structured around learning financial modeling, and credit underwriting',
        10: 'Initiated transactions and managed a portfolio of 10+ companies with revenues from $5 million to $2 billion from various industries including investor real estate, private equity, telecommunications, and agriculture',
        11: 'Partnered with relationship teams to structure and propose credit transactions as well as coordinate customer oversight, documentation, and loan closing processes for new and existing customers '
    }
    bullet_numbers = list(newBulletDict.keys())
    # Find the paragraph that contains the bullet point
    for b in bullet_numbers:
        for para in doc.paragraphs:
            if para.text.strip() == bullet_points[b]:
                para.text = newBulletDict[b]
                for run in para.runs:
                    run.font.name = 'Times'
                    run.font.size = Pt(11)
                    # Ensure the font name is correctly applied in Word
                    r = run._element
                    rPr = r.find(qn('w:rPr'))
                    rFonts = OxmlElement('w:rFonts')
                    rFonts.set(qn('w:ascii'), 'Times')
                    rFonts.set(qn('w:hAnsi'), 'Times')
                    rPr.append(rFonts)
    # Save the document with a new name or overwrite the original
    print('saving doc at : ', doc_path)
    doc.save(doc_path)

# This function takes the json file and returns a list of the urls for all the jobs posted today.
def getURLCompanyTitleAndLocationList(todayJSON):
    urls, companies, titles, locations = [], [], [], []
    with open(todayJSON, 'r') as f:
        data = json.load(f)
        for entry in data:
            link = entry.get("Link")
            company = entry.get("Company")
            title = entry.get("Job Title")
            location = entry.get("Location")
            if link:
                urls.append(link)
            else:
                urls.append(None)
            if company:
                companies.append(company)
            else:
                companies.append(None)
            if title:
                titles.append(title)
            else:
                titles.append(None)
            if location:
                locations.append(location)
            else:
                locations.append(None)
    return [urls, companies, titles, locations]



# This function will be used to create an excel table with the following columns : company name, job title, location, link
# We will take in either the json file or the dict we created in refineJson.
def createXlsxTable(tableName):
    wb = Workbook()
    ws = wb.active
    # Set the headers
    ws['A1'] = 'Company name'
    ws['B1'] = 'Job Title'
    ws['C1'] = 'Location'
    ws['D1'] = 'Indeed Link'
    wb.save(f"{tableName}.xlsx")


# This function will add a job entry
def add_job_entry(tableName, company, job_title, location, link):
    # Load the existing workbook
    wb = load_workbook(f"{tableName}.xlsx")
    ws = wb.active
    # Find the next empty row
    next_row = ws.max_row + 1
    # Insert the data
    ws.cell(row=next_row, column=1, value=company)
    ws.cell(row=next_row, column=2, value=job_title)
    ws.cell(row=next_row, column=3, value=location)
    ws.cell(row=next_row, column=4, value=link)
    wb.save(f"{tableName}.xlsx")

# This will create the folder where we will add all the resumes and spreadsheet. 
def create_folder_if_not_exists(folder_path):
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print(f"Folder '{folder_path}' created.")
    else:
        print(f"Folder '{folder_path}' already exists.")



docx_path = 'Jose Duran Resume.docx'
pdf_path = 'Jose Duran Resume.pdf'
output_pdf_path = 'Jose Duran Resume'
output_docx_path = 'Jose Duran Resume'
todaysDate = f"{datetime.today().strftime('%Y-%m-%d')}"
folderPath = '/Users/keabraekman/Documents-Offline/' + todaysDate


os.makedirs(todaysDate, exist_ok=True)
searchToJson()
jsonPath = todaysDate+'/'+todaysDate+'.json'
refined = refineJson(jsonPath)
print('refined = ', refined)
joburls, companies, titles, locations = getURLCompanyTitleAndLocationList(jsonPath)
createXlsxTable(todaysDate + '/' + todaysDate)

create_folder_if_not_exists(folderPath)

for i in range(len(joburls)):
    if not joburls[i] or not companies[i] or i not in refined:
        continue
    print('working on resume for ', companies[i])
    print('scraping description')
    description = jobDescription(joburls[i])
    print('generating bullets')
    resume_text = generate_resume_bullets(description)
    resume_text_bullets = isolateBullets(resume_text)
    newBulletDict = parse_string_to_dict(resume_text_bullets)
    # title_first_three_words = ' '.join(titles[i].split()[:3])
    title_first_three_words = ' '.join(re.sub(r'\W+', ' ', titles[i]).split()[:3])
    filename = '/Users/keabraekman/Documents-Offline/Jose-Duran/' + todaysDate + '/' + output_docx_path + ' ' + companies[i] + ' ' + title_first_three_words + '.docx'
    folder_path = '/Users/keabraekman/Documents-Offline/Jose-Duran/' + todaysDate + '/'
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    print('replacing bullets')
    replaceBullets(filename, newBulletDict)
    print('adding into spreadsheet')
    add_job_entry(todaysDate+'/'+todaysDate, companies[i], titles[i], locations[i], joburls[i])


driver.quit()

